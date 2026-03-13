"""Utilitarios para integrar OneDrive via Microsoft Graph e pastas locais."""

import os
import re
import base64
import unicodedata
import json
import subprocess
from typing import List, Dict
from urllib.parse import quote
from pathlib import Path

from .graph_client import get_share_info, list_children
from .cache import cached


IMG_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff")
INDEX_EXTENSIONS = IMG_EXTENSIONS + (".lnk",)
SEGMENT_PREFIX_CODE_PATTERN = re.compile(r"^\s*(?P<code>\d{3,8})(?=\D|$)")
GENERIC_CODE_PATTERN = re.compile(r"\b(?P<code>\d{4,8})\b")
PARENTHESIZED_VARIANT_SUFFIX_PATTERN = re.compile(r"\((?P<variant>\d{1,3})\)\s*$")
LEGACY_NUMERIC_VARIANT_PATTERN = re.compile(r"^[-_]\s*(?P<variant>\d{1,3})\s*$")
CATEGORY_STOP_TOKENS = {"C", "COM"}
CATEGORY_TRIM_TAIL_TOKENS = {"C", "COM", "DE", "DA", "DO", "DOS", "DAS", "E"}
STOCK_REPORT_SHEET_NAME = "POSICAO_ESTOQUE"
NO_PHOTO_COVER_URL = "https://placehold.co/900x700?text=Sem+Imagem"
NO_PHOTO_THUMB_URL = "https://placehold.co/240x240?text=Sem+foto"
STOCK_PHOTO_EXTENSIONS = IMG_EXTENSIONS + (".psd",)
FOUR_DIGIT_CODE_PATTERN = re.compile(r"(?<!\d)(?P<code>\d{4})(?!\d)")
DESCRIPTION_MODEL_TOKEN_PATTERN = re.compile(
    r"\b[a-z0-9]+(?:[-_][a-z0-9]+)+\b|\b[a-z]{1,5}\d{2,}[a-z0-9-]*\b",
    re.IGNORECASE,
)
DESCRIPTION_TOKEN_STOPWORDS = {
    "para",
    "com",
    "sem",
    "de",
    "da",
    "do",
    "das",
    "dos",
    "led",
    "bivolt",
    "branco",
    "preto",
    "und",
    "kit",
}


def _encode_share_url(url: str) -> str:
    """Converte uma URL compartilhada do OneDrive no formato Graph "shares/{id}"."""
    clean = url.split("?")[0]
    raw = base64.urlsafe_b64encode(clean.encode()).decode().rstrip("=")
    return f"u!{raw}"


def list_shared_items(share_url: str):
    """Lista os filhos imediatos de uma pasta compartilhada."""
    try:
        info = get_share_info(share_url)
        drive_id = info.get("parentReference", {}).get("driveId")
        item_id = info.get("id")
        if not drive_id or not item_id:
            raise ValueError("unable to resolve shared folder")
        return list_children(drive_id, item_id)
    except Exception as exc:
        raise EnvironmentError(str(exc))


def categorize_photos(items: list, code: str = None) -> dict:
    """Retorna URLs representativas de fotos conforme criterios."""
    result = {"white_background": None, "ambient": None, "measures": None}
    for it in items:
        name = it.get("name", "").lower()
        weburl = it.get("webUrl")
        if code and code.lower() not in name:
            continue
        if result["white_background"] is None and ("branco" in name or "white" in name):
            result["white_background"] = weburl
        if result["ambient"] is None and ("ambient" in name or "ambiente" in name):
            result["ambient"] = weburl
        if result["measures"] is None and ("medida" in name or "measure" in name):
            result["measures"] = weburl
    return result


def _match_filename(name: str, code: str):
    stem = Path(name or "").stem.strip()
    pref = SEGMENT_PREFIX_CODE_PATTERN.match(stem)
    if not pref:
        return None
    if pref.group("code") != str(code):
        return None
    remainder = stem[pref.end() :].strip()
    if not remainder:
        return 0

    legacy_variant = LEGACY_NUMERIC_VARIANT_PATTERN.match(remainder)
    if legacy_variant:
        return int(legacy_variant.group("variant"))

    parenthesized_variant = PARENTHESIZED_VARIANT_SUFFIX_PATTERN.search(remainder)
    if parenthesized_variant:
        return int(parenthesized_variant.group("variant"))

    return 0


def _local_products_paths(base_dir: str) -> List[str]:
    return [
        os.path.join(base_dir, "FOTOS_PRODUTOS"),
        os.path.join(base_dir, "MARKETING", "01_PRODUTOS", "BACKUP PRODUTOS"),
        os.path.join(base_dir, "MARKETING", "Catalogo"),
        os.path.join(base_dir, "MARKETING", "01_PRODUTOS"),
    ]


@cached
def find_images_for_code(share_url: str, code: str, max_depth: int = 5) -> List[Dict]:
    info = get_share_info(share_url)
    drive_id = info.get("parentReference", {}).get("driveId")
    item_id = info.get("id")
    if not drive_id or not item_id:
        raise ValueError("unable to resolve shared folder")
    matches: List[Dict] = []

    def _recurse(d_id: str, i_id: str, depth: int):
        if depth > max_depth:
            return
        children = list_children(d_id, i_id)
        for it in children:
            if "folder" in it:
                _recurse(d_id, it.get("id"), depth + 1)
                continue
            name = it.get("name", "")
            if not any(name.lower().endswith(ext) for ext in IMG_EXTENSIONS):
                continue
            variant = _match_filename(name, code)
            if variant is None:
                continue
            url = it.get("@microsoft.graph.downloadUrl") or (
                f"https://graph.microsoft.com/v1.0/drives/{d_id}/items/{it.get('id')}/content"
            )
            matches.append({"name": name, "variant": variant, "url": url})

    _recurse(drive_id, item_id, 0)
    matches.sort(key=lambda x: x["variant"])
    return matches


def _candidate_local_roots() -> List[str]:
    candidates: List[str] = []
    explicit = os.getenv("CATALOG_LOCAL_PRODUCTS_PATH")
    if explicit:
        candidates.append(explicit)
    for base_env in ("OneDrive", "OneDriveCommercial", "OneDriveConsumer", "USERPROFILE"):
        base = os.getenv(base_env)
        if not base:
            continue
        base = base.rstrip("\\/")
        if base_env == "USERPROFILE":
            one_drive_base = os.path.join(base, "OneDrive")
            candidates.extend(_local_products_paths(one_drive_base))
        else:
            candidates.extend(_local_products_paths(base))
    # Ultima alternativa quando variaveis de ambiente faltam no processo do servidor.
    home = str(Path.home())
    if home:
        candidates.extend(_local_products_paths(os.path.join(home, "OneDrive")))
    return candidates


def _existing_local_roots(path_override: str | None = None) -> List[str]:
    roots: List[str] = []
    seen = set()
    candidates = [path_override] if path_override else _candidate_local_roots()
    for candidate in candidates:
        if not candidate:
            continue
        abs_candidate = os.path.abspath(candidate)
        norm_key = abs_candidate.lower()
        if norm_key in seen:
            continue
        if os.path.isdir(abs_candidate):
            roots.append(abs_candidate)
            seen.add(norm_key)
    return roots


def resolve_local_products_root(path_override: str | None = None) -> str | None:
    roots = _existing_local_roots(path_override)
    return roots[0] if roots else None


def _clean_product_name(candidate: str) -> str:
    cleaned = re.sub(r"\s+", " ", (candidate or "").strip(" -_"))
    cleaned = re.sub(r"\s*\(\d{1,3}\)\s*$", "", cleaned)
    cleaned = re.sub(r"\s*-\s*(atalho|shortcut)\s*$", "", cleaned, flags=re.IGNORECASE)
    cleaned = cleaned.strip(" -_")
    if not cleaned:
        return ""
    # Mantem apenas nomes descritivos; sufixos numericos simples sao variantes.
    if re.fullmatch(r"[1-4]", cleaned):
        return ""
    if re.fullmatch(r"\d+", cleaned):
        return ""
    return cleaned


def _derive_category_from_product_name(product_name: str) -> str:
    cleaned = _clean_product_name(product_name)
    if not cleaned:
        return "Sem categoria"

    tokens = cleaned.split()
    category_tokens: List[str] = []
    for token in tokens:
        token_upper = token.upper()
        if any(char.isdigit() for char in token):
            break
        if token_upper in CATEGORY_STOP_TOKENS and len(category_tokens) >= 2:
            break
        category_tokens.append(token)
        if len(category_tokens) >= 4:
            break

    while category_tokens and category_tokens[-1].upper() in CATEGORY_TRIM_TAIL_TOKENS:
        category_tokens.pop()

    if not category_tokens:
        category_tokens = tokens[:2]

    category = " ".join(category_tokens).strip(" -_")
    return category or "Sem categoria"


def _token_to_category(token: str, next_token: str = "") -> str | None:
    token = (token or "").upper()
    next_token = (next_token or "").upper()
    if not token:
        return None

    if token in {"ARANDELA", "ARANDELAS", "ARANDEDA"} or token.startswith("ARANDEL"):
        return "ARANDELA"
    if token.startswith("ABAJUR"):
        return "ABAJUR"
    if token.startswith("PENDENTE"):
        return "PENDENTE"
    if token.startswith("LUSTRE"):
        return "LUSTRE"
    if token.startswith("PAINEL"):
        return "PAINEL"
    if token.startswith("PLAFON"):
        return "PLAFON"
    if token in {"MINI"} and next_token.startswith("TRILHO"):
        return "TRILHO"
    if token.startswith("TRILHO"):
        return "TRILHO"
    if token.startswith("PERFIL"):
        return "PERFIL"
    if token.startswith("SENSOR"):
        return "SENSOR"
    if token.startswith("BOCAL"):
        return "BOCAL"
    if token.startswith("ESPETO"):
        return "ESPETO"
    if token.startswith("BALIZADOR"):
        return "BALIZADOR"
    if token.startswith("ESPELHO"):
        return "ESPELHO"
    if token.startswith("RELE"):
        return "RELE"
    if token.startswith("REFLET") or token.startswith("HOLOFOTE"):
        return "REFLETOR"
    if token.startswith("LAMP"):
        return "LAMPADA"
    if token.startswith("DRIVER") or token == "DRIVE" or token.startswith("FONTE"):
        return "DRIVER/FONTE"
    if token == "FITA":
        return "FITA LED"
    if token.startswith("LUMINARIA") or token in {"LUMI", "LUM", "LUMIN"}:
        return "LUMINARIA"
    return None


def _normalized_tokens(text: str) -> List[str]:
    normalized = _normalize_name_for_match(text).upper()
    normalized = re.sub(r"[^A-Z0-9\s]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized.split() if normalized else []


def _canonical_category(category: str, product_name: str = "") -> str:
    candidates = [category, product_name]
    for candidate in candidates:
        tokens = _normalized_tokens(candidate)
        if not tokens:
            continue

        if "ILUMINACAO" in tokens and "PUBLICA" in tokens:
            return "ILUMINACAO PUBLICA"
        if "PUBLICA" in tokens and ("LUM" in tokens or "LUMIN" in tokens):
            return "ILUMINACAO PUBLICA"

        first = _token_to_category(tokens[0], tokens[1] if len(tokens) > 1 else "")
        if first:
            return first

        for idx, token in enumerate(tokens[1:], start=1):
            mapped = _token_to_category(token, tokens[idx + 1] if idx + 1 < len(tokens) else "")
            if mapped and mapped != "LUMINARIA":
                return mapped

        if any(
            _token_to_category(token, tokens[idx + 1] if idx + 1 < len(tokens) else "")
            == "LUMINARIA"
            for idx, token in enumerate(tokens)
        ):
            return "LUMINARIA"

        if "FITA" in tokens and "LED" in tokens:
            return "FITA LED"

    return "Sem categoria"


def _extract_code_and_name_from_segment(segment: str) -> tuple[str | None, str]:
    pref = SEGMENT_PREFIX_CODE_PATTERN.search(segment or "")
    if not pref:
        return None, ""

    code = pref.group("code")
    remainder = (segment[pref.end() :] if segment else "").strip()
    remainder = re.sub(r"^[\s\-_]+", "", remainder)
    return code, _clean_product_name(remainder)


def _extract_code_from_parts(parts: List[str]) -> tuple[str | None, str, str]:
    raw_category = parts[0].strip() if len(parts) > 1 else ""
    if raw_category and not SEGMENT_PREFIX_CODE_PATTERN.search(raw_category):
        category = raw_category
    else:
        category = "Sem categoria"

    # Prioridade: nome do arquivo primeiro (geralmente contem descricao), depois pastas pai.
    filename_stem = Path(parts[-1]).stem if parts else ""
    ordered_segments = [filename_stem]
    ordered_segments.extend(reversed(parts[:-1]))

    for segment in ordered_segments:
        code, product_name = _extract_code_and_name_from_segment(segment)
        if code:
            resolved_category = category
            if resolved_category == "Sem categoria" and product_name:
                resolved_category = _derive_category_from_product_name(product_name)
            resolved_category = _canonical_category(resolved_category, product_name)
            return code, product_name or f"Produto {code}", resolved_category

    for segment in ordered_segments:
        generic = GENERIC_CODE_PATTERN.search(segment)
        if generic:
            code = generic.group("code")
            return code, f"Produto {code}", _canonical_category(category, segment)

    return None, "", _canonical_category(category, "")


def _classify_variant(name: str) -> str:
    lowered = name.lower()
    if "branco" in lowered or "white" in lowered:
        return "white_background"
    if "ambient" in lowered or "ambiente" in lowered or "cena" in lowered:
        return "ambient"
    if "medida" in lowered or "measure" in lowered or "dimens" in lowered:
        return "measures"
    return "other"


def _normalize_name_for_match(name: str) -> str:
    normalized = unicodedata.normalize("NFD", name or "")
    without_accents = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return without_accents.lower()


def _is_description_title(name: str, code: str) -> bool:
    normalized = _normalize_name_for_match(name)
    for marker in ("descricao", "description", "descri"):
        if marker in normalized:
            return True

    code_prefix = str(code).strip()
    if not code_prefix:
        return False
    if not normalized.startswith(code_prefix):
        return False

    remainder = normalized[len(code_prefix) :].lstrip()
    if not remainder.startswith("-"):
        return False
    detail = remainder[1:].strip()
    if not detail:
        return False
    # Considera "5989-2.jpg" e sufixos apenas numericos como nao descritivos.
    return not re.match(r"^\d+([._-]|$)", detail)


def _local_file_sort_key(file_info: Dict, code: str) -> tuple:
    name = file_info.get("name", "")
    normalized = _normalize_name_for_match(name)
    if _is_description_title(name, code):
        variant = _match_filename(name, str(code))
        if variant is not None and variant > 0:
            base_without_variant = re.sub(
                r"\s*\(\d{1,3}\)(?=\.[^.]+$)", "", name or "", flags=re.IGNORECASE
            )
            return (0, _normalize_name_for_match(base_without_variant), variant, normalized)
        return (0, normalized)

    variant = _match_filename(name, str(code))
    if variant is not None:
        return (1, variant, normalized)

    return (2, normalized)


def _asset_url(rel_path: str) -> str:
    normalized = rel_path.replace("\\", "/")
    return f"/catalog/local/asset?path={quote(normalized, safe='')}"


def _normalize_category_label(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", (value or "").strip())
    if cleaned in {"***", "-", "â€”"}:
        return ""
    return cleaned


def _rel_path_in_allowed_roots(abs_path: str, roots: List[str]) -> str | None:
    if not abs_path:
        return None
    candidate = os.path.abspath(abs_path)
    if not os.path.isfile(candidate):
        return None
    for root in roots:
        root_abs = os.path.abspath(root)
        try:
            within_root = os.path.commonpath([root_abs, candidate]) == root_abs
        except ValueError:
            continue
        if within_root:
            return os.path.relpath(candidate, root_abs)
    return None


def _resolve_shortcut_targets(scan_root: str) -> Dict[str, str]:
    """Resolve arquivos .lnk em scan_root para caminhos absolutos de destino."""
    if os.name != "nt":
        return {}
    if not scan_root or not os.path.isdir(scan_root):
        return {}

    safe_root = scan_root.replace("'", "''")
    script = (
        f"$root = '{safe_root}'; "
        "$shell = New-Object -ComObject WScript.Shell; "
        "Get-ChildItem -Path $root -Recurse -File -Filter *.lnk | ForEach-Object { "
        "  $target = ''; "
        "  try { $target = $shell.CreateShortcut($_.FullName).TargetPath } catch {} "
        "  [PSCustomObject]@{ link = $_.FullName; target = $target } "
        "} | ConvertTo-Json -Compress"
    )

    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", script],
            capture_output=True,
            text=True,
            check=False,
            timeout=30,
        )
    except Exception:
        return {}

    if result.returncode != 0:
        return {}

    payload = (result.stdout or "").strip()
    if not payload:
        return {}

    try:
        decoded = json.loads(payload)
    except Exception:
        return {}

    rows = decoded if isinstance(decoded, list) else [decoded]
    mapping: Dict[str, str] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        link = str(row.get("link") or "")
        target = str(row.get("target") or "")
        if link and target:
            mapping[os.path.abspath(link)] = os.path.abspath(target)
    return mapping


def _pick_distinct_fallback(files: List[Dict], taken: set[str]) -> Dict | None:
    for item in files:
        rel_path = item.get("rel_path")
        if rel_path and rel_path not in taken:
            return item
    return None


def _scan_local_photo_index(root: str) -> Dict[str, Dict]:
    index: Dict[str, Dict] = {}
    allowed_roots = _existing_local_roots()
    root_abs = os.path.abspath(root)
    if root_abs not in [os.path.abspath(item) for item in allowed_roots]:
        allowed_roots = [root_abs, *allowed_roots]
    shortcut_targets = _resolve_shortcut_targets(root_abs)

    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:
            lowered_name = filename.lower()
            if not lowered_name.endswith(INDEX_EXTENSIONS):
                continue
            full_path = os.path.join(dirpath, filename)
            rel_path = os.path.relpath(full_path, root)
            parts = rel_path.split(os.sep)
            code, product_name, category = _extract_code_from_parts(parts)
            if not code:
                continue
            record = index.setdefault(
                code,
                {
                    "code": code,
                    "name": product_name or f"Produto {code}",
                    "category": category or "Sem categoria",
                    "files": [],
                    "variants": {"white_background": None, "ambient": None, "measures": None},
                },
            )
            if product_name and record["name"].startswith("Produto "):
                record["name"] = product_name
            if category and record["category"] == "Sem categoria":
                record["category"] = category

            if lowered_name.endswith(IMG_EXTENSIONS):
                file_info = {"name": filename, "full_path": full_path, "rel_path": rel_path}
                record["files"].append(file_info)
                variant = _classify_variant(filename)
                if variant in record["variants"] and record["variants"][variant] is None:
                    record["variants"][variant] = file_info
            elif lowered_name.endswith(".lnk"):
                target_path = shortcut_targets.get(os.path.abspath(full_path))
                target_ext = os.path.splitext(target_path or "")[1].lower()
                if target_ext not in IMG_EXTENSIONS:
                    continue
                target_rel_path = _rel_path_in_allowed_roots(target_path, allowed_roots)
                if not target_rel_path:
                    continue
                link_name = filename[:-4] if filename.lower().endswith(".lnk") else filename
                file_info = {
                    "name": link_name,
                    "full_path": target_path,
                    "rel_path": target_rel_path,
                }
                record["files"].append(file_info)
                variant = _classify_variant(link_name)
                if variant in record["variants"] and record["variants"][variant] is None:
                    record["variants"][variant] = file_info

    for record in index.values():
        record["files"].sort(key=lambda item: _local_file_sort_key(item, record.get("code", "")))
        chosen = set()

        white = record["variants"]["white_background"]
        if white is None and record["files"]:
            white = record["files"][0]
            record["variants"]["white_background"] = white
        if white:
            chosen.add(white["rel_path"])

        ambient = record["variants"]["ambient"]
        if ambient:
            chosen.add(ambient["rel_path"])
        else:
            fallback = _pick_distinct_fallback(record["files"], chosen)
            if fallback:
                record["variants"]["ambient"] = fallback
                chosen.add(fallback["rel_path"])

        measures = record["variants"]["measures"]
        if not measures:
            fallback = _pick_distinct_fallback(record["files"], chosen)
            if fallback:
                record["variants"]["measures"] = fallback

    return index


@cached
def build_local_photo_index(root_path: str | None = None) -> Dict[str, Dict]:
    root = resolve_local_products_root(root_path)
    if not root:
        return {}
    return _scan_local_photo_index(root)


def _get_local_index(path_override: str | None = None) -> Dict[str, Dict]:
    """Carrega o indice local do disco para refletir alteracoes de pasta imediatamente."""
    roots = _existing_local_roots(path_override)
    if not roots:
        return {}

    for root in roots:
        rebuilt = _scan_local_photo_index(root)
        if rebuilt:
            return rebuilt
    return {}


def _code_sort_key(code_value: str) -> tuple[int, int | str]:
    code_text = str(code_value or "").strip()
    if code_text.isdigit():
        return (0, int(code_text))
    return (1, code_text)


def _candidate_stock_report_paths() -> List[str]:
    candidates: List[str] = []
    explicit = os.getenv("CATALOG_STOCK_REPORT_PATH")
    if explicit:
        candidates.append(os.path.abspath(explicit))

    repo_root = Path(__file__).resolve().parents[1]
    reports_dir = repo_root / "reports"
    if reports_dir.is_dir():
        workbooks = sorted(
            reports_dir.glob("*.xlsx"),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        for workbook in workbooks:
            if workbook.name.startswith("~$"):
                continue
            candidates.append(str(workbook.resolve()))

    deduped: List[str] = []
    seen = set()
    for candidate in candidates:
        normalized = os.path.abspath(candidate)
        key = normalized.lower()
        if key in seen:
            continue
        deduped.append(normalized)
        seen.add(key)
    return deduped


def _normalize_stock_code(raw_code: object) -> str | None:
    if raw_code is None or isinstance(raw_code, bool):
        return None

    if isinstance(raw_code, int):
        return str(raw_code)
    if isinstance(raw_code, float):
        if raw_code.is_integer():
            return str(int(raw_code))
        raw_text = str(raw_code).strip()
    else:
        raw_text = str(raw_code).strip()

    if not raw_text:
        return None
    normalized = re.sub(r"\.0+$", "", raw_text)
    direct_match = re.fullmatch(r"\d{3,8}", normalized)
    if direct_match:
        return direct_match.group(0)

    generic_match = GENERIC_CODE_PATTERN.search(normalized)
    if generic_match:
        return generic_match.group("code")
    return None


def _placeholder_urls_for_code(code: str) -> tuple[str, str]:
    safe_code = quote(str(code), safe="")
    cover = f"{NO_PHOTO_COVER_URL}+{safe_code}"
    thumb = f"{NO_PHOTO_THUMB_URL}+{safe_code}"
    return cover, thumb


def _resolve_stock_photos_root(path_override: str | None = None) -> str | None:
    candidates: List[str] = []
    if path_override:
        candidates.append(path_override)

    explicit = os.getenv("CATALOG_STOCK_PHOTOS_ROOT")
    if explicit:
        candidates.append(explicit)

    home = str(Path.home())
    if home:
        candidates.append(os.path.join(home, "OneDrive", "MARKETING", "01_PRODUTOS"))

    for candidate in candidates:
        if not candidate:
            continue
        resolved = os.path.abspath(candidate)
        if os.path.isdir(resolved):
            return resolved
    return None


def _extract_four_digit_code(filename: str) -> str | None:
    stem = Path(filename or "").stem
    matches = [m.group("code") for m in FOUR_DIGIT_CODE_PATTERN.finditer(stem)]
    return matches[0] if matches else None


def _extract_four_digit_codes_from_text(text: str) -> List[str]:
    return [m.group("code") for m in FOUR_DIGIT_CODE_PATTERN.finditer(str(text or ""))]


def _normalize_stock_search_text(text: str) -> str:
    normalized = _normalize_name_for_match(text)
    normalized = re.sub(r"[^a-z0-9\s\-_/]", " ", normalized)
    normalized = re.sub(r"[\-_/]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _stock_search_tokens(text: str) -> set[str]:
    normalized = _normalize_stock_search_text(text)
    return {
        token
        for token in normalized.split()
        if len(token) >= 4 and token not in DESCRIPTION_TOKEN_STOPWORDS
    }


def _stock_model_tokens(text: str) -> set[str]:
    normalized = _normalize_name_for_match(text)
    normalized = re.sub(r"[^a-z0-9\s\-_]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return {
        match.group(0)
        for match in DESCRIPTION_MODEL_TOKEN_PATTERN.finditer(normalized)
        if len(match.group(0)) >= 4
    }


def _build_stock_description_profiles(
    description_by_code: Dict[str, List[str]] | None,
    allowed_codes: set[str] | None,
) -> tuple[Dict[str, Dict[str, set[str]]], Dict[str, set[str]], Dict[str, set[str]]]:
    if not description_by_code:
        return {}, {}, {}

    profiles: Dict[str, Dict[str, set[str]]] = {}
    token_to_codes: Dict[str, set[str]] = {}
    model_to_codes: Dict[str, set[str]] = {}

    for code, descriptions in description_by_code.items():
        normalized_code = str(code or "").strip()
        if not re.fullmatch(r"\d{4}", normalized_code):
            continue
        if allowed_codes is not None and normalized_code not in allowed_codes:
            continue

        values = descriptions if isinstance(descriptions, list) else [descriptions]
        token_set: set[str] = set()
        model_set: set[str] = set()
        for value in values:
            token_set.update(_stock_search_tokens(str(value or "")))
            model_set.update(_stock_model_tokens(str(value or "")))

        if not token_set and not model_set:
            continue

        profiles[normalized_code] = {"tokens": token_set, "models": model_set}

        for token in token_set:
            if len(token) < 5:
                continue
            token_to_codes.setdefault(token, set()).add(normalized_code)
        for model in model_set:
            model_to_codes.setdefault(model, set()).add(normalized_code)

    return profiles, token_to_codes, model_to_codes


def _match_stock_code_by_description(
    search_text: str,
    allowed_codes: set[str] | None,
    profiles: Dict[str, Dict[str, set[str]]],
    token_to_codes: Dict[str, set[str]],
    model_to_codes: Dict[str, set[str]],
) -> str | None:
    if not profiles:
        return None

    text_tokens = _stock_search_tokens(search_text)
    text_models = _stock_model_tokens(search_text)
    if not text_tokens and not text_models:
        return None

    candidates: set[str] = set()
    for model in text_models:
        candidates.update(model_to_codes.get(model, set()))

    if not candidates:
        for token in text_tokens:
            candidates.update(token_to_codes.get(token, set()))

    if allowed_codes is not None:
        candidates = {code for code in candidates if code in allowed_codes}
    if not candidates:
        return None

    best_code = None
    best_score = 0
    tie = False
    for code in candidates:
        profile = profiles.get(code)
        if not profile:
            continue
        model_hits = len(text_models & profile.get("models", set()))
        token_hits = len(text_tokens & profile.get("tokens", set()))
        score = (model_hits * 3) + token_hits
        if model_hits == 0 and token_hits < 3:
            continue
        if score > best_score:
            best_code = code
            best_score = score
            tie = False
        elif score == best_score and score > 0:
            tie = True

    if not best_code or tie:
        return None
    if best_score < 4:
        return None
    return best_code


def _scan_stock_photo_index(
    root: str,
    allowed_codes: set[str] | None = None,
    description_by_code: Dict[str, List[str]] | None = None,
) -> Dict[str, Dict]:
    index: Dict[str, Dict] = {}
    normalized_allowed = (
        {str(code) for code in allowed_codes if re.fullmatch(r"\d{4}", str(code))}
        if allowed_codes
        else None
    )
    profiles, token_to_codes, model_to_codes = _build_stock_description_profiles(
        description_by_code=description_by_code,
        allowed_codes=normalized_allowed,
    )

    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:
            ext = os.path.splitext(filename)[1].lower()
            if ext not in STOCK_PHOTO_EXTENSIONS:
                continue

            full_path = os.path.join(dirpath, filename)
            rel_path = os.path.relpath(full_path, root)
            search_text = f"{rel_path} {filename}"

            code = _extract_four_digit_code(filename)
            if code and normalized_allowed is not None and code not in normalized_allowed:
                code = None

            if not code:
                for candidate in _extract_four_digit_codes_from_text(rel_path):
                    if normalized_allowed is None or candidate in normalized_allowed:
                        code = candidate
                        break

            if not code:
                code = _match_stock_code_by_description(
                    search_text=search_text,
                    allowed_codes=normalized_allowed,
                    profiles=profiles,
                    token_to_codes=token_to_codes,
                    model_to_codes=model_to_codes,
                )
            if not code:
                continue

            record = index.setdefault(
                code,
                {
                    "code": code,
                    "name": f"Produto {code}",
                    "category": "Sem categoria",
                    "files": [],
                    "variants": {"white_background": None, "ambient": None, "measures": None},
                },
            )
            file_info = {"name": filename, "full_path": full_path, "rel_path": rel_path}
            record["files"].append(file_info)

            variant = _classify_variant(filename)
            if variant in record["variants"] and record["variants"][variant] is None:
                record["variants"][variant] = file_info

    for code, record in index.items():
        record["files"].sort(key=lambda item: _local_file_sort_key(item, code))
        chosen = set()

        white = record["variants"]["white_background"]
        if white is None and record["files"]:
            white = record["files"][0]
            record["variants"]["white_background"] = white
        if white:
            chosen.add(white["rel_path"])

        ambient = record["variants"]["ambient"]
        if ambient:
            chosen.add(ambient["rel_path"])
        else:
            fallback = _pick_distinct_fallback(record["files"], chosen)
            if fallback:
                record["variants"]["ambient"] = fallback
                chosen.add(fallback["rel_path"])

        measures = record["variants"]["measures"]
        if not measures:
            fallback = _pick_distinct_fallback(record["files"], chosen)
            if fallback:
                record["variants"]["measures"] = fallback

    return index


def _get_stock_photo_index(
    allowed_codes: set[str] | None = None,
    description_by_code: Dict[str, List[str]] | None = None,
) -> Dict[str, Dict]:
    root = _resolve_stock_photos_root()
    if not root:
        return {}
    return _scan_stock_photo_index(
        root,
        allowed_codes=allowed_codes,
        description_by_code=description_by_code,
    )


def _enrich_stock_products_with_photos(products: List[Dict]) -> List[Dict]:
    codes = {
        str(item.get("Codigo", "")).strip()
        for item in products
        if re.fullmatch(r"\d{4}", str(item.get("Codigo", "")).strip())
    }
    description_by_code: Dict[str, List[str]] = {}
    for item in products:
        code = str(item.get("Codigo", "")).strip()
        if not re.fullmatch(r"\d{4}", code):
            continue
        description = str(item.get("Nome") or item.get("Descricao") or "").strip()
        if not description:
            continue
        description_by_code.setdefault(code, []).append(description)

    photo_index = _get_stock_photo_index(
        codes,
        description_by_code=description_by_code,
    )
    if not photo_index:
        return products

    enriched: List[Dict] = []
    for item in products:
        code = str(item.get("Codigo", "")).strip()
        record = photo_index.get(code)
        if not record:
            enriched.append(item)
            continue

        merged = dict(item)
        variants = record.get("variants", {})
        white = variants.get("white_background")
        ambient = variants.get("ambient")
        measures = variants.get("measures")

        white_url = _asset_url(white["rel_path"]) if white else merged.get("URLFoto", "")
        ambient_url = _asset_url(ambient["rel_path"]) if ambient else merged.get("FotoAmbient", "")
        measures_url = _asset_url(measures["rel_path"]) if measures else merged.get("FotoMedidas", "")

        merged["URLFoto"] = white_url
        merged["FotoBranco"] = white_url
        merged["FotoAmbient"] = ambient_url
        merged["FotoMedidas"] = measures_url
        enriched.append(merged)

    return enriched


def _get_stock_photo_record_for_code(code: str) -> Dict | None:
    normalized_code = str(code or "").strip()
    if not re.fullmatch(r"\d{4}", normalized_code):
        return None
    description_by_code: Dict[str, List[str]] = {}
    for item in _load_products_from_available_stock_report():
        candidate_code = str(item.get("Codigo", "")).strip()
        if candidate_code != normalized_code:
            continue
        description = str(item.get("Nome") or item.get("Descricao") or "").strip()
        if not description:
            continue
        description_by_code.setdefault(candidate_code, []).append(description)

    return _get_stock_photo_index(
        {normalized_code},
        description_by_code=description_by_code,
    ).get(normalized_code)


def _load_products_from_stock_report(report_path: str) -> List[Dict]:
    if not report_path or not os.path.isfile(report_path):
        return []

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        print(f"Unable to import openpyxl for stock report: {exc}")
        return []

    try:
        workbook = load_workbook(report_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"Failed to open stock report '{report_path}': {exc}")
        return []

    try:
        if STOCK_REPORT_SHEET_NAME not in workbook.sheetnames:
            return []

        worksheet = workbook[STOCK_REPORT_SHEET_NAME]
        rows: List[tuple[str, int, Dict]] = []

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            raw_code = row[1] if len(row) > 1 else None
            code = _normalize_stock_code(raw_code)
            if not code:
                continue

            raw_description = row[2] if len(row) > 2 else ""
            description = re.sub(r"\s+", " ", str(raw_description or "")).strip()
            if not description:
                description = f"Produto {code}"

            category = _canonical_category("", description)
            cover_url, thumb_url = _placeholder_urls_for_code(code)
            rows.append(
                (
                    code,
                    row_number,
                    {
                        "Codigo": code,
                        "Nome": description,
                        "Descricao": description,
                        "Categoria": category,
                        "URLFoto": cover_url,
                        "Especificacoes": "",
                        "FotoBranco": thumb_url,
                        "FotoAmbient": thumb_url,
                        "FotoMedidas": thumb_url,
                    },
                )
            )

        rows.sort(key=lambda item: (_code_sort_key(item[0]), item[1]))
        return [item[2] for item in rows]
    finally:
        workbook.close()


def _load_products_from_available_stock_report() -> List[Dict]:
    for candidate in _candidate_stock_report_paths():
        loaded = _load_products_from_stock_report(candidate)
        if loaded:
            return loaded
    return []


def _load_cadastro_records(path_override: str | None = None) -> Dict[str, Dict[str, str]]:
    # O catalogo em execucao usa raizes padrao (path_override=None). Para testes e
    # chamadas ad-hoc com override, os dados de origem ficam previsiveis, exceto
    # quando um caminho explicito de cadastro estiver configurado.
    cadastro_path = os.getenv("CATALOG_CADASTRO_HTML")
    if path_override is not None and not cadastro_path:
        return {}

    try:
        from .cadastro import load_cadastro_index

        return load_cadastro_index(cadastro_path)
    except Exception as exc:
        print(f"Failed to load cadastro index: {exc}")
        return {}


def list_local_products(path_override: str | None = None) -> List[Dict]:
    from .erp_catalog import merge_products_with_erp, sort_products_by_category

    index = _get_local_index(path_override)
    if not index and path_override is None:
        stock_products = _load_products_from_available_stock_report()
        if stock_products:
            enriched_stock_products = _enrich_stock_products_with_photos(stock_products)
            return sort_products_by_category(merge_products_with_erp(enriched_stock_products))

    cadastro_records = _load_cadastro_records(path_override)
    products: List[Dict] = []
    for code, record in sorted(index.items(), key=lambda item: _code_sort_key(item[0])):
        variants = record.get("variants", {})
        white = variants.get("white_background")
        ambient = variants.get("ambient")
        measures = variants.get("measures")

        white_url = _asset_url(white["rel_path"]) if white else ""
        ambient_url = _asset_url(ambient["rel_path"]) if ambient else ""
        measures_url = _asset_url(measures["rel_path"]) if measures else ""

        cadastro = cadastro_records.get(str(code), {})
        merged_name = str(cadastro.get("name") or record.get("name") or f"Produto {code}").strip()
        merged_description = str(cadastro.get("description") or "").strip()
        merged_specs = str(cadastro.get("specs") or "").strip()
        cadastro_category = _normalize_category_label(str(cadastro.get("category") or ""))
        if cadastro_category:
            # Mantem categoria exatamente como definida no CADASTRO.html (coluna A).
            merged_category = cadastro_category
        else:
            merged_category_source = _normalize_category_label(
                str(record.get("category") or "Sem categoria")
            )
            merged_category = _canonical_category(merged_category_source, merged_name)

        products.append(
            {
                "Codigo": code,
                "Nome": merged_name,
                "Descricao": merged_description,
                "Categoria": merged_category,
                "URLFoto": white_url,
                "Especificacoes": merged_specs,
                "FotoBranco": white_url,
                "FotoAmbient": ambient_url,
                "FotoMedidas": measures_url,
            }
        )
    return sort_products_by_category(merge_products_with_erp(products))


def categorize_local_photos(code: str, path_override: str | None = None) -> Dict[str, str | None]:
    record = _get_local_index(path_override).get(str(code))
    if not record:
        record = _get_stock_photo_record_for_code(str(code))
    if not record:
        return {"white_background": None, "ambient": None, "measures": None}
    variants = record.get("variants", {})
    return {
        "white_background": _asset_url(variants["white_background"]["rel_path"])
        if variants.get("white_background")
        else None,
        "ambient": _asset_url(variants["ambient"]["rel_path"]) if variants.get("ambient") else None,
        "measures": _asset_url(variants["measures"]["rel_path"]) if variants.get("measures") else None,
    }


def find_local_images_for_code(code: str, path_override: str | None = None) -> List[Dict]:
    record = _get_local_index(path_override).get(str(code))
    if not record:
        record = _get_stock_photo_record_for_code(str(code))
    if not record:
        return []
    ordered_files = sorted(
        record.get("files", []),
        key=lambda item: _local_file_sort_key(item, str(code)),
    )
    images: List[Dict] = []
    for idx, file_info in enumerate(ordered_files):
        images.append(
            {
                "name": file_info.get("name", ""),
                "variant": idx,
                "url": _asset_url(file_info.get("rel_path", "")),
            }
        )
    return images


def resolve_local_asset_path(rel_path: str, path_override: str | None = None) -> str | None:
    roots = _existing_local_roots(path_override)
    stock_root = _resolve_stock_photos_root(path_override)
    if stock_root and stock_root.lower() not in {root.lower() for root in roots}:
        roots = [stock_root, *roots]
    if not roots or not rel_path:
        return None

    normalized = rel_path.replace("/", os.sep).replace("\\", os.sep)
    for root in roots:
        candidate = os.path.abspath(os.path.join(root, normalized))
        try:
            within_root = os.path.commonpath([root, candidate]) == root
        except ValueError:
            continue
        if within_root and os.path.isfile(candidate):
            return candidate
    return None
