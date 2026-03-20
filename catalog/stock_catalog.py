"""Leitura de planilha de estoque e resolucao de fotos associadas."""

from __future__ import annotations

import logging
import os
from pathlib import Path
import re
from typing import Dict, List
from urllib.parse import quote

from .product_media import (
    _asset_url,
    _canonical_category,
    _classify_variant,
    _code_sort_key,
    _local_file_sort_key,
    _normalize_name_for_match,
    _pick_distinct_fallback,
)


logger = logging.getLogger(__name__)

STOCK_REPORT_SHEET_NAME = "POSICAO_ESTOQUE"
NO_PHOTO_COVER_URL = "https://placehold.co/900x700?text=Sem+Imagem"
NO_PHOTO_THUMB_URL = "https://placehold.co/240x240?text=Sem+foto"
STOCK_PHOTO_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".psd")
FOUR_DIGIT_CODE_PATTERN = re.compile(r"(?<!\d)(?P<code>\d{4})(?!\d)")
GENERIC_CODE_PATTERN = re.compile(r"\b(?P<code>\d{4,8})\b")
DESCRIPTION_MODEL_TOKEN_PATTERN = re.compile(
    r"\b[a-z0-9]+(?:[-_][a-z0-9]+)+\b|\b[a-z]{1,5}\d{2,}[a-z0-9-]*\b",
    re.IGNORECASE,
)
DESCRIPTION_TOKEN_STOPWORDS = {
    "para",
    "com",
    "sem",
    "de",
    "da",
    "do",
    "das",
    "dos",
    "led",
    "bivolt",
    "branco",
    "preto",
    "und",
    "kit",
}


def _env_flag(name: str, default: bool = True) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() not in {"0", "false", "no", "off"}


def _candidate_stock_report_paths() -> List[str]:
    candidates: List[str] = []
    explicit = os.getenv("CATALOG_STOCK_REPORT_PATH")
    if explicit:
        candidates.append(os.path.abspath(explicit))

    repo_root = Path(__file__).resolve().parents[1]
    reports_dir = repo_root / "reports"
    if _env_flag("CATALOG_STOCK_REPORT_AUTO_DISCOVERY", default=True) and reports_dir.is_dir():
        workbooks = sorted(
            reports_dir.glob("*.xlsx"),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        for workbook in workbooks:
            if workbook.name.startswith("~$"):
                continue
            candidates.append(str(workbook.resolve()))

    deduped: List[str] = []
    seen = set()
    for candidate in candidates:
        normalized = os.path.abspath(candidate)
        key = normalized.lower()
        if key in seen:
            continue
        deduped.append(normalized)
        seen.add(key)
    return deduped


def _normalize_stock_code(raw_code: object) -> str | None:
    if raw_code is None or isinstance(raw_code, bool):
        return None

    if isinstance(raw_code, int):
        return str(raw_code)
    if isinstance(raw_code, float):
        if raw_code.is_integer():
            return str(int(raw_code))
        raw_text = str(raw_code).strip()
    else:
        raw_text = str(raw_code).strip()

    if not raw_text:
        return None
    normalized = re.sub(r"\.0+$", "", raw_text)
    direct_match = re.fullmatch(r"\d{3,8}", normalized)
    if direct_match:
        return direct_match.group(0)

    generic_match = GENERIC_CODE_PATTERN.search(normalized)
    if generic_match:
        return generic_match.group("code")
    return None


def _placeholder_urls_for_code(code: str) -> tuple[str, str]:
    safe_code = quote(str(code), safe="")
    cover = f"{NO_PHOTO_COVER_URL}+{safe_code}"
    thumb = f"{NO_PHOTO_THUMB_URL}+{safe_code}"
    return cover, thumb


def _resolve_stock_photos_root(path_override: str | None = None) -> str | None:
    candidates: List[str] = []
    if path_override:
        candidates.append(path_override)

    explicit = os.getenv("CATALOG_STOCK_PHOTOS_ROOT")
    if explicit:
        candidates.append(explicit)

    home = str(Path.home())
    if _env_flag("CATALOG_STOCK_PHOTOS_HOME_FALLBACK", default=True) and home:
        candidates.append(os.path.join(home, "OneDrive", "MARKETING", "01_PRODUTOS"))

    for candidate in candidates:
        if not candidate:
            continue
        resolved = os.path.abspath(candidate)
        if os.path.isdir(resolved):
            return resolved
    return None


def _extract_four_digit_code(filename: str) -> str | None:
    stem = Path(filename or "").stem
    matches = [match.group("code") for match in FOUR_DIGIT_CODE_PATTERN.finditer(stem)]
    return matches[0] if matches else None


def _extract_four_digit_codes_from_text(text: str) -> List[str]:
    return [match.group("code") for match in FOUR_DIGIT_CODE_PATTERN.finditer(str(text or ""))]


def _normalize_stock_search_text(text: str) -> str:
    normalized = _normalize_name_for_match(text)
    normalized = re.sub(r"[^a-z0-9\s\-_/]", " ", normalized)
    normalized = re.sub(r"[\-_/]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _stock_search_tokens(text: str) -> set[str]:
    normalized = _normalize_stock_search_text(text)
    return {
        token
        for token in normalized.split()
        if len(token) >= 4 and token not in DESCRIPTION_TOKEN_STOPWORDS
    }


def _stock_model_tokens(text: str) -> set[str]:
    normalized = _normalize_name_for_match(text)
    normalized = re.sub(r"[^a-z0-9\s\-_]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return {
        match.group(0)
        for match in DESCRIPTION_MODEL_TOKEN_PATTERN.finditer(normalized)
        if len(match.group(0)) >= 4
    }


def _build_stock_description_profiles(
    description_by_code: Dict[str, List[str]] | None,
    allowed_codes: set[str] | None,
) -> tuple[Dict[str, Dict[str, set[str]]], Dict[str, set[str]], Dict[str, set[str]]]:
    if not description_by_code:
        return {}, {}, {}

    profiles: Dict[str, Dict[str, set[str]]] = {}
    token_to_codes: Dict[str, set[str]] = {}
    model_to_codes: Dict[str, set[str]] = {}

    for code, descriptions in description_by_code.items():
        normalized_code = str(code or "").strip()
        if not re.fullmatch(r"\d{4}", normalized_code):
            continue
        if allowed_codes is not None and normalized_code not in allowed_codes:
            continue

        values = descriptions if isinstance(descriptions, list) else [descriptions]
        token_set: set[str] = set()
        model_set: set[str] = set()
        for value in values:
            token_set.update(_stock_search_tokens(str(value or "")))
            model_set.update(_stock_model_tokens(str(value or "")))

        if not token_set and not model_set:
            continue

        profiles[normalized_code] = {"tokens": token_set, "models": model_set}

        for token in token_set:
            if len(token) < 5:
                continue
            token_to_codes.setdefault(token, set()).add(normalized_code)
        for model in model_set:
            model_to_codes.setdefault(model, set()).add(normalized_code)

    return profiles, token_to_codes, model_to_codes


def _match_stock_code_by_description(
    search_text: str,
    allowed_codes: set[str] | None,
    profiles: Dict[str, Dict[str, set[str]]],
    token_to_codes: Dict[str, set[str]],
    model_to_codes: Dict[str, set[str]],
) -> str | None:
    if not profiles:
        return None

    text_tokens = _stock_search_tokens(search_text)
    text_models = _stock_model_tokens(search_text)
    if not text_tokens and not text_models:
        return None

    candidates: set[str] = set()
    for model in text_models:
        candidates.update(model_to_codes.get(model, set()))

    if not candidates:
        for token in text_tokens:
            candidates.update(token_to_codes.get(token, set()))

    if allowed_codes is not None:
        candidates = {code for code in candidates if code in allowed_codes}
    if not candidates:
        return None

    best_code = None
    best_score = 0
    tie = False
    for code in candidates:
        profile = profiles.get(code)
        if not profile:
            continue
        model_hits = len(text_models & profile.get("models", set()))
        token_hits = len(text_tokens & profile.get("tokens", set()))
        score = (model_hits * 3) + token_hits
        if model_hits == 0 and token_hits < 3:
            continue
        if score > best_score:
            best_code = code
            best_score = score
            tie = False
        elif score == best_score and score > 0:
            tie = True

    if not best_code or tie:
        return None
    if best_score < 4:
        return None
    return best_code


def _scan_stock_photo_index(
    root: str,
    allowed_codes: set[str] | None = None,
    description_by_code: Dict[str, List[str]] | None = None,
) -> Dict[str, Dict]:
    index: Dict[str, Dict] = {}
    normalized_allowed = (
        {str(code) for code in allowed_codes if re.fullmatch(r"\d{4}", str(code))}
        if allowed_codes
        else None
    )
    profiles, token_to_codes, model_to_codes = _build_stock_description_profiles(
        description_by_code=description_by_code,
        allowed_codes=normalized_allowed,
    )

    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:
            ext = os.path.splitext(filename)[1].lower()
            if ext not in STOCK_PHOTO_EXTENSIONS:
                continue

            full_path = os.path.join(dirpath, filename)
            rel_path = os.path.relpath(full_path, root)
            search_text = f"{rel_path} {filename}"

            code = _extract_four_digit_code(filename)
            if code and normalized_allowed is not None and code not in normalized_allowed:
                code = None

            if not code:
                for candidate in _extract_four_digit_codes_from_text(rel_path):
                    if normalized_allowed is None or candidate in normalized_allowed:
                        code = candidate
                        break

            if not code:
                code = _match_stock_code_by_description(
                    search_text=search_text,
                    allowed_codes=normalized_allowed,
                    profiles=profiles,
                    token_to_codes=token_to_codes,
                    model_to_codes=model_to_codes,
                )
            if not code:
                continue

            record = index.setdefault(
                code,
                {
                    "code": code,
                    "name": f"Produto {code}",
                    "category": "Sem categoria",
                    "files": [],
                    "variants": {"white_background": None, "ambient": None, "measures": None},
                },
            )
            file_info = {"name": filename, "full_path": full_path, "rel_path": rel_path}
            record["files"].append(file_info)

            variant = _classify_variant(filename)
            if variant in record["variants"] and record["variants"][variant] is None:
                record["variants"][variant] = file_info

    for code, record in index.items():
        record["files"].sort(key=lambda item: _local_file_sort_key(item, code))
        chosen = set()

        white = record["variants"]["white_background"]
        if white is None and record["files"]:
            white = record["files"][0]
            record["variants"]["white_background"] = white
        if white:
            chosen.add(white["rel_path"])

        ambient = record["variants"]["ambient"]
        if ambient:
            chosen.add(ambient["rel_path"])
        else:
            fallback = _pick_distinct_fallback(record["files"], chosen)
            if fallback:
                record["variants"]["ambient"] = fallback
                chosen.add(fallback["rel_path"])

        measures = record["variants"]["measures"]
        if not measures:
            fallback = _pick_distinct_fallback(record["files"], chosen)
            if fallback:
                record["variants"]["measures"] = fallback

    return index


def _get_stock_photo_index(
    allowed_codes: set[str] | None = None,
    description_by_code: Dict[str, List[str]] | None = None,
) -> Dict[str, Dict]:
    root = _resolve_stock_photos_root()
    if not root:
        return {}
    return _scan_stock_photo_index(
        root,
        allowed_codes=allowed_codes,
        description_by_code=description_by_code,
    )


def _enrich_stock_products_with_photos(products: List[Dict]) -> List[Dict]:
    codes = {
        str(item.get("Codigo", "")).strip()
        for item in products
        if re.fullmatch(r"\d{4}", str(item.get("Codigo", "")).strip())
    }
    description_by_code: Dict[str, List[str]] = {}
    for item in products:
        code = str(item.get("Codigo", "")).strip()
        if not re.fullmatch(r"\d{4}", code):
            continue
        description = str(item.get("Nome") or item.get("Descricao") or "").strip()
        if not description:
            continue
        description_by_code.setdefault(code, []).append(description)

    photo_index = _get_stock_photo_index(
        codes,
        description_by_code=description_by_code,
    )
    if not photo_index:
        return products

    enriched: List[Dict] = []
    for item in products:
        code = str(item.get("Codigo", "")).strip()
        record = photo_index.get(code)
        if not record:
            enriched.append(item)
            continue

        merged = dict(item)
        variants = record.get("variants", {})
        white = variants.get("white_background")
        ambient = variants.get("ambient")
        measures = variants.get("measures")

        white_url = _asset_url(white["rel_path"]) if white else merged.get("URLFoto", "")
        ambient_url = _asset_url(ambient["rel_path"]) if ambient else merged.get("FotoAmbient", "")
        measures_url = _asset_url(measures["rel_path"]) if measures else merged.get("FotoMedidas", "")

        merged["URLFoto"] = white_url
        merged["FotoBranco"] = white_url
        merged["FotoAmbient"] = ambient_url
        merged["FotoMedidas"] = measures_url
        enriched.append(merged)

    return enriched


def _get_stock_photo_records_for_codes(codes: set[str]) -> Dict[str, Dict]:
    normalized_codes = {str(code or "").strip() for code in codes if re.fullmatch(r"\d{4}", str(code or "").strip())}
    if not normalized_codes:
        return {}

    description_by_code: Dict[str, List[str]] = {}
    for item in _load_products_from_available_stock_report():
        code = str(item.get("Codigo", "")).strip()
        if code not in normalized_codes:
            continue
        description = str(item.get("Nome") or item.get("Descricao") or "").strip()
        if not description:
            continue
        description_by_code.setdefault(code, []).append(description)

    return _get_stock_photo_index(
        normalized_codes,
        description_by_code=description_by_code,
    )


def _get_stock_photo_record_for_code(code: str) -> Dict | None:
    normalized_code = str(code or "").strip()
    if not re.fullmatch(r"\d{4}", normalized_code):
        return None
    return _get_stock_photo_records_for_codes({normalized_code}).get(normalized_code)


def _load_products_from_stock_report(report_path: str) -> List[Dict]:
    if not report_path or not os.path.isfile(report_path):
        return []

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        logger.warning("Unable to import openpyxl for stock report: %s", exc)
        return []

    try:
        workbook = load_workbook(report_path, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Failed to open stock report '%s': %s", report_path, exc)
        return []

    try:
        if STOCK_REPORT_SHEET_NAME not in workbook.sheetnames:
            return []

        worksheet = workbook[STOCK_REPORT_SHEET_NAME]
        rows: List[tuple[str, int, Dict]] = []

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            raw_code = row[1] if len(row) > 1 else None
            code = _normalize_stock_code(raw_code)
            if not code:
                continue

            raw_description = row[2] if len(row) > 2 else ""
            description = re.sub(r"\s+", " ", str(raw_description or "")).strip()
            if not description:
                description = f"Produto {code}"

            category = _canonical_category("", description)
            cover_url, thumb_url = _placeholder_urls_for_code(code)
            rows.append(
                (
                    code,
                    row_number,
                    {
                        "Codigo": code,
                        "Nome": description,
                        "Descricao": description,
                        "Categoria": category,
                        "URLFoto": cover_url,
                        "Especificacoes": "",
                        "FotoBranco": thumb_url,
                        "FotoAmbient": thumb_url,
                        "FotoMedidas": thumb_url,
                    },
                )
            )

        rows.sort(key=lambda item: (_code_sort_key(item[0]), item[1]))
        return [item[2] for item in rows]
    finally:
        workbook.close()


def _load_products_from_available_stock_report() -> List[Dict]:
    for candidate in _candidate_stock_report_paths():
        loaded = _load_products_from_stock_report(candidate)
        if loaded:
            return loaded
    return []
